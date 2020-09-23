import os
import re
import threading
import datetime
import pathlib
import hashlib
import openpyxl

import tkinter as tk
import tkinter.ttk
import tkinter.filedialog
import tkinter.messagebox

VER = 'Ver 1.0.5'

def find_all_files(dir):
    files = []
    for f in pathlib.Path(dir).rglob('*'):
        files.append(str(f))
    return files

def md5(file):
    hash_md5 = hashlib.md5()

    try:
        f = open(file, 'rb')
        for chunk in iter(lambda: f.read(4096), b''):
            hash_md5.update(chunk)
    except:
        return False

    return hash_md5.hexdigest()

def start(dir, xlsx):
    button03['state'] = tk.DISABLED
    button03.update()
    progress01['value'] = 0

    header = [
        'name',
        'ext',
        'bytes',
        'mtime'
    ]

    chk01_checked = chk01_var.get()
    if chk01_checked:
        header.append('md5')

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'files'
    sheet.append(header)

    fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='FFFFCC')
    for column in range(1, len(header) + 1):
        sheet.cell(row=1, column=column).fill = fill

    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20

    if chk01_checked:
        sheet.column_dimensions['E'].width = 35

    sheet.auto_filter.ref = sheet.dimensions
    sheet.freeze_panes = 'A2'

    try:
        wb.save(xlsx)
    except:
        finish()
        tk.messagebox.showerror('エラー', '結果を書き込めませんでした。')
        return

    try:
        os.chdir(dir)
        files = find_all_files('.')
    except Exception as e:
        finish()
        tk.messagebox.showerror('エラー', e)
        return

    length = len(files)
    for i in range(length):
        file = files[i]

        filename = re.sub(r'^\.[\\/]', '', file)
        filename = openpyxl.cell.cell.ILLEGAL_CHARACTERS_RE.sub('', file)

        if file == '.':
            continue
        elif not os.path.exists(file):
            continue

        ext = ''
        if os.path.isdir(file):
            ext = 'DIR'
        elif os.path.islink(file):
            ext = 'LINK'
        elif '.' in os.path.basename(file):
            ext = os.path.basename(file).split('.')[-1].lower()

        try:
            stat = os.stat(file)
        except:
            sheet.append([filename, ext])
            continue

        data = [
            filename,
            ext,
            stat.st_size,
            datetime.datetime.fromtimestamp(stat.st_mtime)
        ]

        if chk01_checked and os.path.isfile(file):
            md5sum = md5(file)
            data.append(md5sum)

        sheet.append(data)

        sheet.cell(i + 2, 3).number_format = '#,##0'
        sheet.cell(i + 2, 4).number_format = 'yyyy/mm/dd hh:mm:ss'

        if i % 100 == 99:
            progress01['value'] = i * 100 / length

    wb.save(xlsx)

    progress01['value'] = 100
    finish()
    tk.messagebox.showinfo('終了', '完了しました。')

def finish():
    button03['state'] = tk.NORMAL
    button03.update()

def validation_dir(dir):
    if not dir or not os.path.isdir(dir):
        tk.messagebox.showerror('エラー', '基準フォルダを選択してください。')
        return False

    return True

def button01_clicked():
    dir = tk.filedialog.askdirectory()

    if not dir:
        return

    if not validation_dir(dir):
        return

    entry01.delete(0, tk.END)
    entry01.insert(tk.END, dir)

def button02_clicked():
    dir = entry01.get()

    if not validation_dir(dir):
        return

    initf = 'timestamp_' + os.path.splitext(os.path.basename(dir))[0] + '.xlsx'
    if os.name == 'nt':
        ftypes = [('Excel Books', '.xlsx')]
        filename = tk.filedialog.asksaveasfilename(filetype=ftypes, initialfile=initf)
    else:
        filename = tk.filedialog.asksaveasfilename(initialfile=initf)

    entry02.delete(0, tk.END)
    entry02.insert(tk.END, filename)

def button03_clicked():
    dir = entry01.get()
    xlsx = entry02.get()

    if not validation_dir(dir):
        return

    if not xlsx:
        tk.messagebox.showerror('エラー', '結果保存先を指定してください。')
        return

    thread = threading.Thread(target=start, args=([dir, xlsx]))
    thread.start()

if __name__ == '__main__':
    root = tk.Tk()
    root.title('タイムスタンプ調べるくん ' + VER)

    frame01 = tk.Frame(root, padx=20, pady=10)
    frame01.pack()

    row = 0
    label01 = tk.Label(frame01, text='基準フォルダ:', font=('', 14))
    label01.grid(column=0, row=row, pady=10, sticky='w')
    entry01 = tk.Entry(frame01, width=40)
    entry01.grid(column=1, row=row, pady=10)
    button01 = tk.Button(frame01, text='参照', command=button01_clicked, font=('', 14))
    button01.grid(column=2, row=row, pady=10)

    row += 1
    label02 = tk.Label(frame01, text='結果保存先:', font=('', 14))
    label02.grid(column=0, row=row, pady=10, sticky='w')
    entry02 = tk.Entry(frame01, width=40)
    entry02.grid(column=1, row=row, pady=10)
    button02 = tk.Button(frame01, text='参照', command=button02_clicked, font=('', 14))
    button02.grid(column=2, row=row, pady=10)

    row += 1
    chk01_var = tk.IntVar()
    chk01 = tk.Checkbutton(frame01, text='ハッシュ値 (md5) を計算', variable=chk01_var)
    chk01.grid(column=0, row=row, columnspan=3, pady=10, sticky='w')

    row += 1
    progress01 = tk.ttk.Progressbar(frame01, orient=tk.HORIZONTAL, length=300, mode='determinate')
    progress01.grid(column=0, row=row, columnspan=3, pady=10)

    row += 1
    button03 = tk.Button(frame01, text='実　　行', command=button03_clicked, width=20, font=('', 20))
    button03.grid(column=0, row=row, columnspan=3, pady=10)

    root.mainloop()